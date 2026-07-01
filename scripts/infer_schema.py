#!/usr/bin/env python3
"""Infer simple schema information from discovered source files.

This script reads the discovery report created by scan_source.py and looks at a
small sample of each supported file. It then writes a beginner-friendly JSON
summary with column names and basic data types.
"""

from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, List
import zipfile
import xml.etree.ElementTree as ET


def get_project_root() -> Path:
    """Return the project root even when this script lives in the scripts folder."""
    current_path = Path(__file__).resolve()
    if current_path.parent.name == "scripts":
        return current_path.parent.parent
    return current_path.parent


# These are the file types we want to inspect.
SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".json", ".parquet", ".xml", ".txt"}


def infer_data_type(values: List[Any]) -> str:
    """Return a simple data type for a list of values.

    The logic is intentionally simple so it is easy to follow for beginners.
    We check for the most common cases in this order:
    1. DATETIME
    2. INTEGER
    3. DECIMAL
    4. BOOLEAN
    5. STRING
    """
    cleaned_values = [value for value in values if value is not None and str(value).strip() != ""]

    if not cleaned_values:
        return "STRING"

    # Try DATETIME first. This catches values like 2026-01-01.
    datetime_patterns = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
    ]
    if all(_looks_like_datetime(str(value), datetime_patterns) for value in cleaned_values):
        return "DATETIME"

    # Try INTEGER if every value is a whole number.
    if all(_looks_like_integer(str(value)) for value in cleaned_values):
        return "INTEGER"

    # Try DECIMAL if every value looks like a number with a decimal point.
    if all(_looks_like_decimal(str(value)) for value in cleaned_values):
        return "DECIMAL"

    # Try BOOLEAN for simple true/false values.
    if all(_looks_like_boolean(str(value)) for value in cleaned_values):
        return "BOOLEAN"

    return "STRING"


def _looks_like_datetime(value: str, patterns: List[str]) -> bool:
    """Return True if the given value can be parsed as a date/time."""
    for pattern in patterns:
        try:
            datetime.strptime(value, pattern)
            return True
        except ValueError:
            continue
    return False


def _looks_like_integer(value: str) -> bool:
    """Return True if the value is a whole number."""
    return bool(re.fullmatch(r"[+-]?\d+", value))


def _looks_like_decimal(value: str) -> bool:
    """Return True if the value is a number with decimals."""
    return bool(re.fullmatch(r"[+-]?\d+\.\d+", value))


def _looks_like_boolean(value: str) -> bool:
    """Return True if the value is a simple true/false style value."""
    return value.strip().lower() in {"true", "false", "yes", "no", "1", "0"}


def infer_from_csv(file_path: Path, row_limit: int = 100) -> List[dict]:
    """Inspect the first rows of a CSV file and infer column information."""
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)

    if not rows:
        return []

    # The first row is treated as the header row.
    headers = [str(cell).strip() or f"column_{index}" for index, cell in enumerate(rows[0])]
    data_rows = rows[1:row_limit]

    # Build a list of values for each column.
    column_values: dict[str, List[Any]] = {header: [] for header in headers}
    for row in data_rows:
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else ""
            column_values[header].append(value)

    return [
        {"name": header, "data_type": infer_data_type(column_values[header])}
        for header in headers
    ]


def infer_from_json(file_path: Path, record_limit: int = 100) -> List[dict]:
    """Inspect the first records from a JSON file and infer column information."""
    with file_path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)

    if isinstance(data, list):
        records = data[:record_limit]
    elif isinstance(data, dict):
        records = [data]
    else:
        records = []

    if not records:
        return []

    # Collect column names from dictionaries.
    column_names: List[str] = []
    for record in records:
        if isinstance(record, dict):
            for key in record.keys():
                if key not in column_names:
                    column_names.append(str(key))

    column_values: dict[str, List[Any]] = {name: [] for name in column_names}
    for record in records:
        if isinstance(record, dict):
            for name in column_names:
                column_values[name].append(record.get(name))

    return [
        {"name": name, "data_type": infer_data_type(column_values[name])}
        for name in column_names
    ]


def infer_from_xlsx(file_path: Path, row_limit: int = 100) -> List[dict]:
    """Inspect the first sheet of an Excel workbook.

    This uses openpyxl if it is available. If it is not installed, we return an
    empty result and record the error message in the output file.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed") from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [str(cell).strip() or f"column_{index}" for index, cell in enumerate(rows[0])]
    data_rows = rows[1:row_limit]

    column_values: dict[str, List[Any]] = {header: [] for header in headers}
    for row in data_rows:
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else ""
            column_values[header].append(value)

    return [
        {"name": header, "data_type": infer_data_type(column_values[header])}
        for header in headers
    ]


def infer_from_xls(file_path: Path, row_limit: int = 100) -> List[dict]:
    """Inspect the first sheet of an XLS file.

    This is more difficult than XLSX and may require an extra package. We handle
    the situation gracefully by raising an error if the needed library is missing.
    """
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("pandas is not installed for XLS processing") from exc

    dataframe = pd.read_excel(file_path, sheet_name=0, nrows=row_limit)
    if dataframe.empty:
        return []

    headers = [str(header).strip() or f"column_{index}" for index, header in enumerate(dataframe.columns)]
    columns = []
    for header in headers:
        values = dataframe[header].tolist() if header in dataframe.columns else []
        columns.append({"name": header, "data_type": infer_data_type(values)})
    return columns


def resolve_source_file(source_root: Path, dataset_name: str, file_name: str) -> Path:
    """Find the actual source file that matches the discovery entry."""
    dataset_dir = source_root / dataset_name
    if not dataset_dir.exists():
        raise FileNotFoundError(f"Dataset folder not found: {dataset_dir}")

    matches = list(dataset_dir.rglob(file_name))
    if matches:
        return matches[0]

    raise FileNotFoundError(f"File not found: {dataset_name}/{file_name}")


def build_schema_report(discovery_path: Path, source_root: Path) -> dict:
    """Create the final JSON report from the discovery summary."""
    if not discovery_path.exists():
        return {
            "source_discovery_file": str(discovery_path),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "files": [],
            "errors": ["Discovery file was not found."],
        }

    with discovery_path.open("r", encoding="utf-8") as handle:
        discovery_data = json.load(handle)

    files = []
    errors = []

    for item in discovery_data.get("discovered_files", []):
        dataset_name = item.get("dataset_name", "")
        file_name = item.get("file_name", "")
        extension = item.get("extension", "")

        try:
            source_file = resolve_source_file(source_root, dataset_name, file_name)
            columns = []

            if extension.lower() == "csv":
                columns = infer_from_csv(source_file)
            elif extension.lower() in {"xlsx"}:
                columns = infer_from_xlsx(source_file)
            elif extension.lower() in {"xls"}:
                columns = infer_from_xls(source_file)
            elif extension.lower() == "json":
                columns = infer_from_json(source_file)
            else:
                # Unsupported file types are simply recorded as not inspected.
                columns = []

            files.append(
                {
                    "dataset_name": dataset_name,
                    "file_name": file_name,
                    "extension": extension,
                    "columns": columns,
                }
            )
        except Exception as exc:  # noqa: BLE001
            errors.append({
                "dataset_name": dataset_name,
                "file_name": file_name,
                "extension": extension,
                "error": str(exc),
            })

    return {
        "source_discovery_file": str(discovery_path),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "files": files,
        "errors": errors,
    }


def main() -> None:
    """Run the schema inference process and save the output JSON file."""
    root_dir = get_project_root()
    discovery_path = root_dir / "outputs" / "discovered_datasets.json"
    source_root = root_dir / "source"
    output_path = root_dir / "outputs" / "inferred_schema.json"

    report = build_schema_report(discovery_path, source_root)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        json.dump(report, handle, indent=2)

    print(f"Schema report saved to {output_path}")


if __name__ == "__main__":
    main()
